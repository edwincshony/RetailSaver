from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.mixins import LoginRequiredMixin
from django.contrib import messages
from django.db.models import Q
from django.core.paginator import Paginator
from django.views.generic import ListView, CreateView, UpdateView, DeleteView, View
from utils.pagination import paginate_queryset  # make sure path is correct
from django.urls import reverse_lazy
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from .models import Product
from .forms import ProductForm
from reportlab.lib.pagesizes import letter
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from datetime import datetime

class AdminRequiredMixin(LoginRequiredMixin):
    def dispatch(self, request, *args, **kwargs):
        if not request.user.is_authenticated:
            messages.warning(request, 'You must be logged in to access this page.')
            return HttpResponseRedirect(reverse_lazy('login'))
        if not hasattr(request.user, 'role') or request.user.role != 'admin':
            messages.warning(request, 'No permission to access this page.')
            return HttpResponseRedirect(reverse_lazy('login'))
        return super().dispatch(request, *args, **kwargs)

from django.http import JsonResponse
from django.db.models import Q

class ProductListView(AdminRequiredMixin, ListView):
    model = Product
    template_name = 'products/product_list.html'
    context_object_name = 'products'
    paginate_by = None

    def get_queryset(self):
        queryset = Product.objects.filter(user=self.request.user)
        search = self.request.GET.get('search')
        if search:
            queryset = queryset.filter(Q(name__icontains=search))
        return queryset.order_by('-date_added')

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        page_obj, products = paginate_queryset(self.request, context['object_list'])
        context['products'] = products
        context['page_obj'] = page_obj
        context['paginator'] = page_obj.paginator
        context['is_paginated'] = True
        context['search_term'] = self.request.GET.get('search', '')
        return context

    def get(self, request, *args, **kwargs):
        # Check if it's an AJAX request
        if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
            self.object_list = self.get_queryset()
            context = self.get_context_data()
            
            # Return JSON data instead of HTML
            products_data = []
            for product in context['products']:
                products_data.append({
                    'id': product.pk,
                    'name': product.name,
                    'weight': product.get_weight_display(),
                    'amount': product.amount,
                    'date_added': product.date_added.strftime('%b %d, %Y %H:%M'),
                    'update_url': request.build_absolute_uri(f'/products/{product.pk}/update/'),
                    'delete_url': request.build_absolute_uri(f'/products/{product.pk}/delete/'),
                })
            
            return JsonResponse({'products': products_data})
        
        # Regular request
        return super().get(request, *args, **kwargs)




class ProductCreateView(AdminRequiredMixin, CreateView):
    model = Product
    form_class = ProductForm
    template_name = 'products/product_form.html'
    success_url = reverse_lazy('product_list')

    def form_valid(self, form):
        form.instance.user = self.request.user
        messages.success(self.request, 'Product added successfully.')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Form error. Please correct the fields.')
        return super().form_invalid(form)

class ProductUpdateView(AdminRequiredMixin, UpdateView):
    model = Product
    form_class = ProductForm
    template_name = 'products/product_form.html'
    success_url = reverse_lazy('product_list')

    def get_queryset(self):
        # Only allow edit if owner or superuser
        return Product.objects.filter(user=self.request.user)
    
    def form_valid(self, form):
        messages.success(self.request, 'Product updated successfully.')
        return super().form_valid(form)
    
    def form_invalid(self, form):
        messages.error(self.request, 'Form error. Please correct the fields.')
        return super().form_invalid(form)

class ProductDeleteView(AdminRequiredMixin, DeleteView):
    model = Product
    template_name = 'products/product_confirm_delete.html'
    success_url = reverse_lazy('product_list')

    def get_queryset(self):
        # Only allow delete if owner or superuser
        return Product.objects.filter(user=self.request.user)
    
    def delete(self, request, *args, **kwargs):
        messages.success(self.request, 'Product deleted successfully.')
        return super().delete(request, *args, **kwargs)


class ProductExportPDFView(AdminRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        products = Product.objects.filter(user=request.user).order_by('-date_added')
        
        # Create PDF
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
        
        doc = SimpleDocTemplate(response, pagesize=letter)
        elements = []
        
        # Add title
        styles = getSampleStyleSheet()
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=16,
            textColor=colors.HexColor('#1f77b4'),
            spaceAfter=20,
            alignment=1
        )
        elements.append(Paragraph("AFC Bank Supermarket Products", title_style))
        elements.append(Spacer(1, 12))
        
        # Create table data
        data = [['Product Name', 'Quantity', 'Amount (Rs.)', 'Date Added']]
        for product in products:
            data.append([
                product.name,
                product.get_weight_display(),
                f"{product.amount}",
                product.date_added.strftime('%b %d, %Y')
            ])
        
        # Create table
        table = Table(data, colWidths=[150, 100, 100, 120])
        table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#1f77b4')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#f0f0f0')])
        ]))
        
        elements.append(table)
        elements.append(Spacer(1, 20))
        
        # Add summary
        summary_style = ParagraphStyle(
            'Summary',
            parent=styles['Normal'],
            fontSize=10,
            textColor=colors.grey,
        )
        total_amount = sum(float(p.amount) for p in products)
        elements.append(Paragraph(f"<b>Total Products:</b> {products.count()}", summary_style))
        elements.append(Paragraph(f"<b>Total Value:</b> Rs. {total_amount:.2f}", summary_style))
        elements.append(Paragraph(f"<b>Generated on:</b> {datetime.now().strftime('%B %d, %Y %I:%M %p')}", summary_style))
        
        doc.build(elements)
        return response


class ProductExportExcelView(AdminRequiredMixin, View):
    def get(self, request, *args, **kwargs):
        products = Product.objects.filter(user=request.user).order_by('-date_added')
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        
        # Add headers
        headers = ['Product Name', 'Quantity', 'Amount (Rs.)', 'Date Added']
        ws.append(headers)
        
        # Style headers
        header_fill = PatternFill(start_color='1f77b4', end_color='1f77b4', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=12)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center', vertical='center')
        
        # Add data
        for product in products:
            ws.append([
                product.name,
                product.get_weight_display(),
                float(product.amount),
                product.date_added.strftime('%b %d, %Y')
            ])
        
        # Format columns
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        
        # Add summary section
        summary_row = len(products) + 3
        ws[f'A{summary_row}'] = 'Total Products:'
        ws[f'B{summary_row}'] = products.count()
        
        ws[f'A{summary_row + 1}'] = 'Total Value (Rs.):'
        total_amount = sum(float(p.amount) for p in products)
        ws[f'B{summary_row + 1}'] = total_amount
        
        # Make summary bold
        for row in ws[summary_row:summary_row + 2]:
            for cell in row:
                cell.font = Font(bold=True)
        
        # Return as download
        response = HttpResponse(
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="products_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
        wb.save(response)
        return response